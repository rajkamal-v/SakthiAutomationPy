import csv
import os

parent_dir = os.getcwd()

def get_csv_data(filename):
    filepath = os.path.abspath(parent_dir + "/data/" + filename)
    data = []
    with open(filepath) as file:
        obj = csv.reader(file)
        row_key = next(obj)
        for row in obj:
            data.append({row_key[i]:row[i] for i in range(len(row))})
        return data


from openpyxl import load_workbook

def get_excel_data(filename, sheetname):
    filepath = os.path.abspath(parent_dir + "/data/" + filename)
    data = []
    wb = load_workbook(filepath)
    sheet = wb[sheetname]
    rows = sheet.max_row
    cols = sheet.max_column

    for row in range(2,rows+1):
        row_dict = dict()
        for col in range(1, cols + 1):
            key = sheet.cell(row=1,column=col).value
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value == None:
                cell_value = ""
            row_dict.setdefault(key,cell_value)
        data.append(row_dict)

    return data

